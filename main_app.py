import sys
import os
import cv2
import numpy as np
import torch
import torchvision.transforms as transforms
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QLabel, QFileDialog, QComboBox, QCheckBox,
                             QMessageBox, QDialog, QSizePolicy, QFrame, QScrollArea,
                             QGridLayout, QButtonGroup, QRadioButton, QSlider)
from PyQt5.QtGui import QImage, QPixmap, QFont
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QDateTime, pyqtSlot
import traceback
import time
from collections import Counter
import shutil

import torch.nn as nn
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class SimpleCNN(nn.Module):
    def __init__(self, num_classes=7):
        super(SimpleCNN, self).__init__()

        self.conv1 = nn.Conv2d(3, 32, kernel_size=6, padding=1)
        self.conv2 = nn.Conv2d(32, 64, kernel_size=6, padding=1)

        self.fc1 = nn.Linear(10816, 512)
        self.fc2 = nn.Linear(512, 128)
        self.fc3 = nn.Linear(128, num_classes)

        self.pool = nn.MaxPool2d(kernel_size=6, stride=3, padding=0)

        self.relu = nn.LeakyReLU()

        self.dropout = nn.Dropout(0.3)

    def forward(self, x):
        x = self.relu(self.conv1(x))
        x = self.pool(self.relu(self.conv2(x)))

        x = x.view(-1, 10816)

        x = self.relu(self.fc1(x))
        x = self.dropout(x)
        x = self.relu(self.fc2(x))
        x = self.dropout(x)
        x = self.fc3(x)

        return x

EMOTION_DICT_ENG = {
    0: "Angry", 1: "Disgust", 2: "Fear", 3: "Happy",
    4: "Sad", 5: "Surprise", 6: "Neutral"
}
EMOTION_DICT_RU = {
    0: "Злость", 1: "Отвращение", 2: "Страх", 3: "Счастье",
    4: "Грусть", 5: "Удивление", 6: "Нейтральность"
}
EMOTION_COLORS = ['#ff9999','#66b3ff','#99ff99','#ffcc99','#c2c2f0','#ffb3e6', '#c9c9c9']

try:
    from model import SimpleCNN
except ImportError:
    import torch.nn as nn
    class SimpleCNN(nn.Module):
        def __init__(self, num_classes=7): super().__init__(); self.fc = nn.Linear(1, num_classes)
        def forward(self, x): return self.fc(torch.randn(x.size(0), 1))


class VideoThread(QThread):
    change_pixmap_signal = pyqtSignal(np.ndarray)
    stop_frame_signal = pyqtSignal(np.ndarray, str)
    auto_stop_frame_detected_signal = pyqtSignal(np.ndarray, str, str, int)
    processing_finished_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)
    new_emotion_data_for_plot_signal = pyqtSignal(float, int)
    video_duration_ready_signal = pyqtSignal(int)
    current_video_position_signal = pyqtSignal(int)
    clear_plot_signal = pyqtSignal()

    def __init__(self, source_type, source_path, model_path, target_emotion_idx, manual_stop_frame_mode, device):
        super().__init__()
        self.source_type = source_type
        self.source_path = source_path
        self.model_path = model_path
        self.target_emotion_idx = target_emotion_idx
        self.manual_stop_frame_mode = manual_stop_frame_mode
        self.device = device
        self._run_flag = True
        self._pause_flag = False
        self.model = None
        self.face_cascade = None
        self.transform = None
        self.current_stop_frame_pending_user_action = False
        self.emotion_log = []
        self.start_time_s = 0
        self.cap = None
        self.seek_to_msec = -1

    def load_resources(self):
        try:
            self.model = SimpleCNN(num_classes=len(EMOTION_DICT_ENG))
            self.model.load_state_dict(torch.load(self.model_path, map_location=self.device))
            self.model.to(self.device)
            self.model.eval()

            haar_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
            if not os.path.exists(haar_path):
                base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
                haar_path_local = os.path.join(base_path, 'haarcascade_frontalface_default.xml')
                if os.path.exists(haar_path_local):
                    haar_path = haar_path_local
                else:
                    self.error_signal.emit("Файл haarcascade_frontalface_default.xml не найден.")
                    return False

            self.face_cascade = cv2.CascadeClassifier(haar_path)
            if self.face_cascade.empty():
                self.error_signal.emit("Не удалось загрузить каскад Хаара.")
                return False

            self.transform = transforms.Compose([
                transforms.ToPILImage(),
                transforms.Resize((48, 48)),
                transforms.ToTensor(),
                transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
            ])
            return True
        except Exception as e:
            self.error_signal.emit(f"Ошибка загрузки ресурсов: {e}\n{traceback.format_exc()}")
            return False

    def run(self):
        if not self.load_resources():
            self.processing_finished_signal.emit(self.emotion_log)
            return

        cap_args = [int(self.source_path) if self.source_type == "webcam" else self.source_path]
        if self.source_type == "webcam" and sys.platform == "win32":
             cap_args.append(cv2.CAP_DSHOW)
        self.cap = cv2.VideoCapture(*cap_args)

        if not self.cap.isOpened():
            self.error_signal.emit(f"Не удалось открыть источник: {self.source_path}")
            self.processing_finished_signal.emit(self.emotion_log)
            return

        fps = 30
        if self.source_type == "video":
            duration_ms = 0
            total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            video_fps = self.cap.get(cv2.CAP_PROP_FPS)
            if video_fps > 0 and total_frames > 0:
                duration_ms = int((total_frames / video_fps) * 1000)
                fps = video_fps
            self.video_duration_ready_signal.emit(duration_ms)

        self.start_time_s = time.time()
        frame_count = 0

        while self._run_flag and self.cap.isOpened():
            while self._pause_flag and self._run_flag:
                self.msleep(100)
            if not self._run_flag:
                break

            if self.current_stop_frame_pending_user_action:
                self.msleep(100)
                continue

            if self.seek_to_msec >= 0 and self.source_type == "video":
                seek_success = self.cap.set(cv2.CAP_PROP_POS_MSEC, self.seek_to_msec)
                if seek_success:
                    self.emotion_log = []
                    self.clear_plot_signal.emit()
                    current_frame_num_after_seek = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
                    frame_count = int(current_frame_num_after_seek) if current_frame_num_after_seek is not None else 0
                else:
                    self.error_signal.emit(f"Не удалось перемотать видео на {self.seek_to_msec} мс.")
                self.seek_to_msec = -1

            ret, frame = self.cap.read()
            if not ret:
                break

            current_time_s = (time.time() - self.start_time_s) if self.source_type == "webcam" \
                             else self.cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0

            if self.source_type == "video":
                self.current_video_position_signal.emit(int(self.cap.get(cv2.CAP_PROP_POS_MSEC)))

            frame_for_display = frame.copy()
            gray_for_detection = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray_for_detection, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            # Находим ближайшее лицо (самое большое по площади, т.е. к камере)
            if len(faces) > 0:
                # Сортируем лица по площади (w * h) в порядке убывания
                faces = sorted(faces, key=lambda x: x[2] * x[3], reverse=True)
                # Берем только первое (самое большое) лицо
                x, y, w, h = faces[0]
                
                face_roi_color_bgr = frame[y:y+h, x:x+w]
                try:
                    img_tensor = self.transform(face_roi_color_bgr).unsqueeze(0).to(self.device)
                    with torch.no_grad():
                        outputs = self.model(img_tensor)
                        probabilities = torch.softmax(outputs, dim=1)
                        confidence_tensor, predicted_idx_tensor = torch.max(probabilities, 1)

                    predicted_emotion_idx = predicted_idx_tensor.item()
                    emotion_name_eng = EMOTION_DICT_ENG.get(predicted_emotion_idx, "Unknown")
                    emotion_name_ru = EMOTION_DICT_RU.get(predicted_emotion_idx, "Неизвестно")
                    confidence = confidence_tensor.item() * 100

                    self.emotion_log.append({
                        'time_s': round(current_time_s, 2), 'frame': frame_count,
                        'face_idx_on_frame': 0, 'emotion_idx': predicted_emotion_idx,
                        'emotion_name_eng': emotion_name_eng, 'emotion_name_ru': emotion_name_ru,
                        'confidence': round(confidence, 2), 'bbox': (x,y,w,h)
                    })

                    self.new_emotion_data_for_plot_signal.emit(current_time_s, predicted_emotion_idx)

                    cv2.rectangle(frame_for_display, (x, y), (x+w, y+h), (0, 255, 0), 2)
                    label_for_video = f"{emotion_name_eng}: {confidence:.1f}%"
                    cv2.putText(frame_for_display, label_for_video, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

                    if predicted_emotion_idx == self.target_emotion_idx:
                        if self.manual_stop_frame_mode:
                            self.stop_frame_signal.emit(frame.copy(), emotion_name_ru)
                            self.current_stop_frame_pending_user_action = True
                        else:
                            frame_copy_for_auto = frame.copy()
                            timestamp_auto = QDateTime.currentMSecsSinceEpoch()
                            self.auto_stop_frame_detected_signal.emit(frame_copy_for_auto, emotion_name_ru, emotion_name_eng, timestamp_auto)
                except Exception:
                    pass

            if not self.current_stop_frame_pending_user_action:
                self.change_pixmap_signal.emit(frame_for_display)
                frame_count += 1

            processing_time_ms = 15
            delay_ms = max(1, int(1000 / fps) - processing_time_ms)
            self.msleep(delay_ms)

        if self.cap:
            self.cap.release()
        self.processing_finished_signal.emit(self.emotion_log)

    def resume_processing_manual(self):
        self.current_stop_frame_pending_user_action = False

    def pause_stream(self):
        self._pause_flag = True

    def resume_stream(self):
        self._pause_flag = False

    def stop_stream(self):
        self._run_flag = False
        self._pause_flag = False
        self.wait()

    def seek_video(self, msec_position):
        if self.source_type == "video" and self.cap:
            self.seek_to_msec = msec_position

class StopFrameDialog(QDialog):
    StopProcessingAction = QDialog.Accepted + 1

    def __init__(self, frame_np, emotion_name_ru, parent_app_ref, parent=None):
        super().__init__(parent)
        self.parent_app = parent_app_ref
        self.setWindowTitle(f"Обнаружен стоп-кадр: {emotion_name_ru}")
        self.frame_to_save = frame_np.copy()
        layout = QVBoxLayout(self)
        self.image_label = QLabel(); self.image_label.setAlignment(Qt.AlignCenter)
        q_img = self.convert_cv_qt(frame_np)
        self.image_label.setPixmap(q_img.scaled(640, 480, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        layout.addWidget(self.image_label)
        top_button_layout = QHBoxLayout()
        self.save_button = QPushButton("Сохранить")
        self.discard_button = QPushButton("Пропустить")
        top_button_layout.addWidget(self.save_button); top_button_layout.addWidget(self.discard_button)
        layout.addLayout(top_button_layout)
        self.stop_processing_button = QPushButton("Стоп (завершить обработку)")
        layout.addWidget(self.stop_processing_button)
        self.save_button.clicked.connect(self.accept)
        self.discard_button.clicked.connect(self.reject)
        self.stop_processing_button.clicked.connect(self.trigger_stop_processing)
    def trigger_stop_processing(self):
        if self.parent_app:
            self.parent_app.stop_processing()
        self.done(StopFrameDialog.StopProcessingAction)
    def convert_cv_qt(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape; bytes_per_line = ch * w
        return QPixmap.fromImage(QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888))
    def get_frame_to_save(self): return self.frame_to_save

class ClickableLabel(QLabel):
    leftClicked = pyqtSignal(str)

    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.image_path = image_path
        self.setCursor(Qt.PointingHandCursor)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.leftClicked.emit(self.image_path)
        super().mousePressEvent(event)

class EnlargedThumbnailViewer(QDialog):
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Просмотр: {os.path.basename(image_path)}")
        
        self.original_pixmap = QPixmap(image_path)

        self.resize(800, 600) 
        self.setMinimumSize(400, 300)

        layout = QVBoxLayout(self)
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        layout.addWidget(self.image_label)

        if self.original_pixmap.isNull():
            self.image_label.setText("Не удалось загрузить изображение.")
        
        button_box = QHBoxLayout()
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        button_box.addWidget(spacer)
        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(self.accept)
        button_box.addWidget(close_button)
        layout.addLayout(button_box)
        
        if not self.original_pixmap.isNull():
             self.update_pixmap()

    def update_pixmap(self):
        if not self.original_pixmap.isNull() and self.image_label.width() > 0 and self.image_label.height() > 0 :
            scaled_pixmap = self.original_pixmap.scaled(
                self.image_label.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.image_label.setPixmap(scaled_pixmap)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_pixmap()

class EmotionDetectorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Детектор Эмоций")
        self.setGeometry(50, 50, 1400, 900)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QHBoxLayout(self.central_widget)

        self.settings_widget = QWidget()
        self.settings_widget.setMaximumWidth(320)
        self.settings_layout = QVBoxLayout(self.settings_widget)

        self.model_path_label = QLabel("Модель (.pth): Не выбрана")
        self.settings_layout.addWidget(self.model_path_label)
        self.load_model_button = QPushButton("Загрузить модель")
        self.load_model_button.clicked.connect(self.load_model)
        self.settings_layout.addWidget(self.load_model_button)

        self.settings_layout.addWidget(QLabel("Источник данных:"))
        self.webcam_radio = QRadioButton("Веб-камера")
        self.video_file_radio = QRadioButton("Видеофайл")
        self.source_button_group = QButtonGroup(self)
        self.source_button_group.addButton(self.webcam_radio)
        self.source_button_group.addButton(self.video_file_radio)
        self.webcam_radio.setChecked(True)
        self.webcam_radio.toggled.connect(self.source_changed)
        self.settings_layout.addWidget(self.webcam_radio)
        self.webcam_combo = QComboBox()
        self.populate_webcams()
        self.settings_layout.addWidget(self.webcam_combo)
        self.settings_layout.addWidget(self.video_file_radio)
        self.video_path_label = QLabel("Видеофайл: Не выбран")
        self.settings_layout.addWidget(self.video_path_label)
        self.load_video_button = QPushButton("Загрузить видео")
        self.load_video_button.clicked.connect(self.load_video)
        self.load_video_button.setEnabled(False)
        self.settings_layout.addWidget(self.load_video_button)

        self.settings_layout.addWidget(QLabel("Ловить стоп-кадр на эмоцию:"))
        self.emotion_combo = QComboBox()
        for idx, name_ru in EMOTION_DICT_RU.items():
            self.emotion_combo.addItem(name_ru, idx)
        self.settings_layout.addWidget(self.emotion_combo)
        self.manual_mode_checkbox = QCheckBox("Включить ручной режим стоп-кадров")
        self.manual_mode_checkbox.setChecked(False)
        self.settings_layout.addWidget(self.manual_mode_checkbox)

        self.start_button = QPushButton("Старт")
        self.start_button.clicked.connect(self.start_processing)
        self.settings_layout.addWidget(self.start_button)
        self.pause_button = QPushButton("Пауза")
        self.pause_button.clicked.connect(self.toggle_pause_resume)
        self.pause_button.setEnabled(False)
        self.settings_layout.addWidget(self.pause_button)
        self.stop_button = QPushButton("Стоп")
        self.stop_button.clicked.connect(self.stop_processing)
        self.stop_button.setEnabled(False)
        self.settings_layout.addWidget(self.stop_button)

        self.save_all_frames_button = QPushButton("Сохранить все найденные стоп-кадры")
        self.save_all_frames_button.clicked.connect(self.save_all_collected_frames)
        self.save_all_frames_button.setEnabled(False)
        self.settings_layout.addWidget(self.save_all_frames_button)

        self.save_report_button = QPushButton("Сохранить отчет (.xlsx)")
        self.save_report_button.clicked.connect(self.save_report_xlsx)
        self.save_report_button.setEnabled(False)
        if openpyxl is None:
            self.save_report_button.setToolTip("Библиотека openpyxl не найдена, сохранение в XLSX недоступно.")
        self.settings_layout.addWidget(self.save_report_button)

        self.create_pie_chart_button = QPushButton("Создать диаграмму эмоций")
        self.create_pie_chart_button.clicked.connect(self.create_emotion_pie_chart)
        self.create_pie_chart_button.setEnabled(False)
        self.settings_layout.addWidget(self.create_pie_chart_button)

        self.settings_layout.addStretch()
        self.main_layout.addWidget(self.settings_widget, 1)

        self.center_visualization_area = QWidget()
        self.center_visualization_layout = QVBoxLayout(self.center_visualization_area)

        self.live_video_label = QLabel("Ожидание запуска...")
        self.live_video_label.setAlignment(Qt.AlignCenter)
        self.live_video_label.setMinimumSize(640, 360)
        self.live_video_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.live_video_label.setStyleSheet("border: 1px solid black; background-color: #f0f0f0;")
        self.center_visualization_layout.addWidget(self.live_video_label, 3)

        self.video_slider = QSlider(Qt.Horizontal)
        self.video_slider.setMinimumHeight(25)
        self.video_slider.setVisible(False)
        self.video_slider.setEnabled(False)
        self.video_slider.sliderPressed.connect(self.slider_pressed)
        self.video_slider.sliderReleased.connect(self.slider_released)
        self.video_slider.valueChanged.connect(self.update_time_label_from_slider)
        self.center_visualization_layout.addWidget(self.video_slider)

        self.time_display_label = QLabel("00:00 / 00:00")
        self.time_display_label.setAlignment(Qt.AlignCenter)
        self.time_display_label.setVisible(False)
        self.center_visualization_layout.addWidget(self.time_display_label)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        self.center_visualization_layout.addWidget(line)

        self.plot_and_stopframe_container = QWidget()
        self.plot_and_stopframe_layout = QHBoxLayout(self.plot_and_stopframe_container)

        self.plot_canvas_widget = QWidget()
        plot_layout = QVBoxLayout(self.plot_canvas_widget)
        self.figure = Figure(figsize=(5, 3), dpi=100)
        self.plot_canvas = FigureCanvas(self.figure)
        plot_layout.addWidget(self.plot_canvas)
        self.ax_plot = self.figure.add_subplot(111)
        self.init_plot()
        self.plot_canvas_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.plot_and_stopframe_layout.addWidget(self.plot_canvas_widget, 2)

        self.stop_frame_display_label = QLabel("Стоп-кадр")
        self.stop_frame_display_label.setAlignment(Qt.AlignCenter)
        self.stop_frame_display_label.setMinimumSize(160,120)
        self.stop_frame_display_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.stop_frame_display_label.setStyleSheet("border: 1px solid gray; background-color: #e0e0e0;")
        self.plot_and_stopframe_layout.addWidget(self.stop_frame_display_label, 1)

        self.center_visualization_layout.addWidget(self.plot_and_stopframe_container, 2)

        self.main_layout.addWidget(self.center_visualization_area, 3)

        self.thumbnails_scroll_area = QScrollArea()
        self.thumbnails_scroll_area.setWidgetResizable(True)
        self.thumbnails_scroll_area.setMinimumWidth(200)
        self.thumbnails_scroll_area.setMaximumWidth(220)
        self.thumbnails_widget = QWidget()
        self.thumbnails_layout = QVBoxLayout(self.thumbnails_widget)
        self.thumbnails_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        self.thumbnails_scroll_area.setWidget(self.thumbnails_widget)
        self.main_layout.addWidget(self.thumbnails_scroll_area, 0)

        self.model_path = None
        self.video_path = None
        self.video_thread = None
        self.is_paused = False
        self.collected_emotion_data_for_report = []
        self.collected_auto_stop_frames_data = []
        self.plot_times = []
        self.plot_emotions_idx = []
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

        self.stop_frame_clear_timer = QTimer(self)
        self.stop_frame_clear_timer.setSingleShot(True)
        self.stop_frame_clear_timer.timeout.connect(lambda: self.stop_frame_display_label.setText("Стоп-кадр"))
        
        self.temp_auto_thumbs_dir = os.path.join("results", "temp_auto_thumbnails")
        self.manual_saved_thumb_paths = []

        self.user_is_scrubbing = False
        self.current_video_duration_ms = 0
        self.total_duration_str = "00:00"

    def format_ms_to_time(self, ms):
        if ms < 0: ms = 0
        s = round(ms / 1000)
        m = s // 60
        s = s % 60
        return f"{m:02d}:{s:02d}"

    def init_plot(self):
        self.ax_plot.clear()
        self.ax_plot.set_title("Преобладающая эмоция", fontsize=9)
        self.ax_plot.set_xlabel("Время (секунды)", fontsize=8)
        self.ax_plot.set_ylabel("Эмоция", fontsize=8)
        self.ax_plot.set_yticks(list(EMOTION_DICT_RU.keys()))
        self.ax_plot.set_yticklabels(list(EMOTION_DICT_RU.values()), fontdict={'fontsize': 7})
        self.ax_plot.grid(True, linestyle='--', alpha=0.7)
        self.plot_line, = self.ax_plot.plot([], [], 'b-o', markersize=3, linewidth=1.5)
        self.figure.subplots_adjust(left=0.28, bottom=0.20, right=0.95, top=0.88)
        self.plot_canvas.draw()
        self.plot_times = []
        self.plot_emotions_idx = []

    @pyqtSlot(float, int)
    def update_plot_live(self, time_s, emotion_idx):
        self.plot_times.append(time_s)
        self.plot_emotions_idx.append(emotion_idx)

        max_points_on_plot = 100
        if len(self.plot_times) > max_points_on_plot:
            self.plot_times = self.plot_times[-max_points_on_plot:]
            self.plot_emotions_idx = self.plot_emotions_idx[-max_points_on_plot:]

        self.plot_line.set_data(self.plot_times, self.plot_emotions_idx)

        if self.plot_times:
            min_time = self.plot_times[0]
            max_time = self.plot_times[-1]
            self.ax_plot.set_xlim(min_time, max_time + max(1, (max_time - min_time) * 0.05))

            if max_time - min_time > 10:
                self.ax_plot.xaxis.set_major_locator(plt.MaxNLocator(nbins=5, integer=True))
            else:
                self.ax_plot.xaxis.set_major_locator(plt.MaxNLocator(integer=True))

        self.ax_plot.relim()
        self.ax_plot.autoscale_view(True,True,True)
        self.ax_plot.set_yticks(list(EMOTION_DICT_RU.keys()))
        self.ax_plot.set_yticklabels(list(EMOTION_DICT_RU.values()), fontdict={'fontsize': 7})
        self.plot_canvas.draw_idle()

    def populate_webcams(self):
        self.webcam_combo.clear()
        available_cameras = []
        for i in range(5):
            cap_test_args = [i]
            if sys.platform == "win32":
                cap_test_args.append(cv2.CAP_DSHOW)
            cap_test = cv2.VideoCapture(*cap_test_args)
            if cap_test.isOpened():
                available_cameras.append(f"Камера {i}")
            cap_test.release()

        if not available_cameras:
            self.webcam_combo.addItem("Нет доступных камер", -1)
            self.webcam_combo.setEnabled(False)
        else:
            for name in available_cameras:
                cam_real_idx = int(name.split()[-1])
                self.webcam_combo.addItem(name, cam_real_idx)
            self.webcam_combo.setEnabled(True)

    def load_model(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выбрать файл модели", "", "PyTorch Model (*.pth)")
        if path:
            self.model_path = path
            self.model_path_label.setText(f"Модель: ...{os.path.basename(path)[-30:]}")
        else:
            self.model_path = None
            self.model_path_label.setText("Модель (.pth): Не выбрана")

    def source_changed(self):
        is_webcam = self.webcam_radio.isChecked()
        self.webcam_combo.setEnabled(is_webcam)
        self.load_video_button.setEnabled(not is_webcam)

        self.video_slider.setVisible(not is_webcam)
        self.time_display_label.setVisible(not is_webcam)
        if is_webcam:
            self.video_path_label.setText("Видеофайл: Не выбран (веб-камера)")
            self.video_slider.setEnabled(False)
        else:
            self.video_path_label.setText("Видеофайл: Не выбран" if not self.video_path else f"Видеофайл: ...{os.path.basename(self.video_path)[-30:]}")
            self.video_slider.setEnabled(self.current_video_duration_ms > 0 and not (self.video_thread and self.video_thread.isRunning()))

    def load_video(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выбрать видеофайл", "", "Video Files (*.mp4 *.avi *.mkv *.mov)")
        if path:
            self.video_path = path
            self.video_path_label.setText(f"Видеофайл: ...{os.path.basename(path)[-30:]}")

    def start_processing(self):
        if not self.model_path:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл модели.")
            return

        source_type = "webcam" if self.webcam_radio.isChecked() else "video"
        source_path = None
        if source_type == "webcam":
            if self.webcam_combo.currentData() == -1 or self.webcam_combo.currentData() is None:
                QMessageBox.warning(self, "Ошибка", "Веб-камера не выбрана или не доступна.")
                return
            source_path = str(self.webcam_combo.currentData())
        else:
            if not self.video_path or not os.path.exists(self.video_path):
                QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите корректный видеофайл.")
                return
            source_path = self.video_path

        self.clear_thumbnails()
        self.manual_saved_thumb_paths = []
        self.collected_auto_stop_frames_data = []

        if os.path.exists(self.temp_auto_thumbs_dir):
            shutil.rmtree(self.temp_auto_thumbs_dir)
        os.makedirs(self.temp_auto_thumbs_dir, exist_ok=True)

        self.set_controls_enabled_on_start(False)

        self.live_video_label.setText("Запуск обработки...")
        self.stop_frame_display_label.setText("Стоп-кадр")
        self.stop_frame_display_label.clear()
        self.collected_emotion_data_for_report = []
        self.init_plot()

        self.time_display_label.setText("00:00 / 00:00")
        if source_type == "video":
            self.video_slider.setValue(0)
            self.video_slider.setEnabled(True)
            self.video_slider.setVisible(True)
            self.time_display_label.setVisible(True)
        else:
            self.video_slider.setEnabled(False)
            self.video_slider.setVisible(False)
            self.time_display_label.setVisible(False)
            self.current_video_duration_ms = 0
            self.total_duration_str = "00:00"

        manual_mode_active = self.manual_mode_checkbox.isChecked()
        self.video_thread = VideoThread(source_type, source_path, self.model_path,
                                        self.emotion_combo.currentData(), manual_mode_active, self.device)

        self.video_thread.change_pixmap_signal.connect(self.update_live_video)
        self.video_thread.stop_frame_signal.connect(self.handle_manual_stop_frame_dialog)
        self.video_thread.auto_stop_frame_detected_signal.connect(self.handle_auto_stop_frame_and_thumbnail)
        self.video_thread.processing_finished_signal.connect(self.processing_finished)
        self.video_thread.error_signal.connect(self.show_error_message)
        self.video_thread.new_emotion_data_for_plot_signal.connect(self.update_plot_live)
        
        self.video_thread.video_duration_ready_signal.connect(self.set_video_duration)
        self.video_thread.current_video_position_signal.connect(self.update_slider_position)
        self.video_thread.clear_plot_signal.connect(self.handle_clear_plot_request)

        self.video_thread.start()

    def toggle_pause_resume(self):
        if self.video_thread and self.video_thread.isRunning():
            self.is_paused = not self.is_paused
            if self.is_paused:
                self.video_thread.pause_stream()
                self.pause_button.setText("Продолжить")
            else:
                self.video_thread.resume_stream()
                self.pause_button.setText("Пауза")

            can_generate_reports = (not self.is_paused or not (self.video_thread and self.video_thread.isRunning())) and \
                                   bool(self.collected_emotion_data_for_report)
            self.save_report_button.setEnabled(openpyxl is not None and can_generate_reports)
            self.create_pie_chart_button.setEnabled(can_generate_reports)

            can_save_all_frames = (not self.is_paused or not (self.video_thread and self.video_thread.isRunning())) and \
                                  not self.manual_mode_checkbox.isChecked() and \
                                  bool(self.collected_auto_stop_frames_data)
            self.save_all_frames_button.setEnabled(can_save_all_frames)

    def stop_processing(self):
        if self.video_thread and self.video_thread.isRunning():
            self.live_video_label.setText("Остановка...")
            self.video_thread.stop_stream()

    def set_controls_enabled_on_start(self, enabled):
        self.load_model_button.setEnabled(enabled)
        self.webcam_radio.setEnabled(enabled)
        self.video_file_radio.setEnabled(enabled)

        if enabled:
            self.source_changed()
        else:
            self.webcam_combo.setEnabled(False)
            self.load_video_button.setEnabled(False)

        self.emotion_combo.setEnabled(enabled)
        self.manual_mode_checkbox.setEnabled(enabled)
        self.start_button.setEnabled(enabled)

        self.stop_button.setEnabled(not enabled)
        self.pause_button.setEnabled(not enabled)

        is_video_mode = self.video_file_radio.isChecked()

        if enabled:
            self.pause_button.setText("Пауза")
            self.is_paused = False

            has_data_for_report = bool(self.collected_emotion_data_for_report)
            self.save_report_button.setEnabled(openpyxl is not None and has_data_for_report)
            self.create_pie_chart_button.setEnabled(has_data_for_report)

            has_auto_stop_frames = bool(self.collected_auto_stop_frames_data)
            self.save_all_frames_button.setEnabled(not self.manual_mode_checkbox.isChecked() and has_auto_stop_frames)

            self.video_slider.setVisible(is_video_mode)
            self.time_display_label.setVisible(is_video_mode)
            self.video_slider.setEnabled(is_video_mode and self.current_video_duration_ms > 0)
        else:
            self.save_report_button.setEnabled(False)
            self.create_pie_chart_button.setEnabled(False)
            self.save_all_frames_button.setEnabled(False)

            self.video_slider.setVisible(is_video_mode)
            self.time_display_label.setVisible(is_video_mode)
            self.video_slider.setEnabled(is_video_mode)

    @pyqtSlot(np.ndarray)
    def update_live_video(self, cv_img):
        qt_img = self.convert_cv_qt(cv_img)
        scaled_pixmap = qt_img.scaled(self.live_video_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.live_video_label.setPixmap(scaled_pixmap)

    @pyqtSlot(np.ndarray, str)
    def handle_manual_stop_frame_dialog(self, frame_np, emotion_name_ru):
        if not (self.manual_mode_checkbox.isChecked() and self.video_thread):
            return

        qt_img_small = self.convert_cv_qt(frame_np)
        self.stop_frame_display_label.setPixmap(qt_img_small.scaled(self.stop_frame_display_label.size(),
                                                                    Qt.KeepAspectRatio, Qt.SmoothTransformation))
        dialog = StopFrameDialog(frame_np, emotion_name_ru, self, self)
        dialog_result = dialog.exec_()

        if dialog_result == QDialog.Accepted:
            filepath = self.save_manual_frame(dialog.get_frame_to_save(), emotion_name_ru)
            if filepath:
                self.add_thumbnail(filepath, is_manual=True)
            if self.video_thread: self.video_thread.resume_processing_manual()
        elif dialog_result == QDialog.Rejected:
            if self.video_thread: self.video_thread.resume_processing_manual()
        elif dialog_result == StopFrameDialog.StopProcessingAction:
            if self.video_thread: self.video_thread.resume_processing_manual()

        self.stop_frame_clear_timer.start(3000)

    @pyqtSlot(np.ndarray, str, str, int)
    def handle_auto_stop_frame_and_thumbnail(self, frame_np, emotion_name_ru, emotion_name_eng, timestamp_ms):
        if self.manual_mode_checkbox.isChecked():
            return

        qt_img_stop_frame = self.convert_cv_qt(frame_np)
        self.stop_frame_display_label.setPixmap(qt_img_stop_frame.scaled(
            self.stop_frame_display_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))

        temp_thumb_filename = f"auto_thumb_{timestamp_ms}_{emotion_name_eng}.png"
        temp_thumb_path = os.path.join(self.temp_auto_thumbs_dir, temp_thumb_filename)
        try:
            cv2.imwrite(temp_thumb_path, frame_np)
            self.add_thumbnail(temp_thumb_path, is_manual=False)
        except Exception as e:
            print(f"Error saving temp auto thumbnail: {e}")
            return

        self.collected_auto_stop_frames_data.append({
            'frame': frame_np.copy(),
            'emotion_name_ru': emotion_name_ru,
            'emotion_name_eng': emotion_name_eng,
            'timestamp': timestamp_ms,
            'temp_path': temp_thumb_path
        })
        self.save_all_frames_button.setEnabled(True)


    def save_manual_frame(self, frame_np, emotion_name_ru):
        base_results_dir = "results"
        os.makedirs(base_results_dir, exist_ok=True)
        emotion_idx = next((idx for idx, name in EMOTION_DICT_RU.items() if name == emotion_name_ru), None)
        emotion_name_eng_for_folder = EMOTION_DICT_ENG.get(emotion_idx, "Unknown") if emotion_idx is not None else "Unknown"
        emotion_folder = os.path.join(base_results_dir, emotion_name_eng_for_folder)
        os.makedirs(emotion_folder, exist_ok=True)
        timestamp = QDateTime.currentDateTime().toString("yyyyMMdd_HHmmss_zzz")
        filename = os.path.join(emotion_folder, f"manual_frame_{timestamp}.png")
        try:
            cv2.imwrite(filename, frame_np)
            QMessageBox.information(self, "Сохранено", f"Кадр сохранен как {filename}")
            return filename
        except Exception as e:
            QMessageBox.critical(self, "Ошибка сохранения", f"Не удалось сохранить кадр: {e}")
            return None

    @pyqtSlot(list)
    def processing_finished(self, emotion_log_from_thread):
        self.live_video_label.setText("Обработка завершена или остановлена.")
        self.collected_emotion_data_for_report = emotion_log_from_thread

        self.set_controls_enabled_on_start(True)

        if self.video_file_radio.isChecked() and self.current_video_duration_ms > 0:
            if self.collected_emotion_data_for_report:
                is_stopped_by_user = self.video_thread is not None and not self.video_thread.isRunning() and self.video_thread._run_flag == False
                if not is_stopped_by_user and self.video_slider.value() < self.video_slider.maximum():
                     pass
        self.video_thread = None
        if self.collected_emotion_data_for_report:
            if openpyxl: QMessageBox.information(self, "Отчет", "Данные для отчета XLSX собраны.")
            QMessageBox.information(self, "Диаграмма", "Данные для круговой диаграммы собраны.")
        if not self.manual_mode_checkbox.isChecked() and self.collected_auto_stop_frames_data:
            QMessageBox.information(self, "Стоп-кадры", f"Собрано {len(self.collected_auto_stop_frames_data)} автоматических стоп-кадров для сохранения.")

    def save_all_collected_frames(self):
        if not self.collected_auto_stop_frames_data:
            QMessageBox.information(self, "Нет кадров", "Нет собранных стоп-кадров для сохранения.")
            return

        base_folder_path = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения стоп-кадров", "results")
        if base_folder_path:
            saved_count = 0
            for i, frame_data in enumerate(self.collected_auto_stop_frames_data):
                frame_np = frame_data['frame']
                emotion_name_eng = frame_data['emotion_name_eng']
                timestamp_ms = frame_data.get('timestamp', QDateTime.currentMSecsSinceEpoch() + i)
                q_datetime = QDateTime.fromMSecsSinceEpoch(timestamp_ms)
                time_str = q_datetime.toString("yyyyMMdd_HHmmss_zzz")
                emotion_folder_path = os.path.join(base_folder_path, emotion_name_eng)
                os.makedirs(emotion_folder_path, exist_ok=True)
                filename = os.path.join(emotion_folder_path, f"auto_frame_{time_str}.png")
                try:
                    cv2.imwrite(filename, frame_np)
                    saved_count += 1
                except Exception as e:
                    print(f"Ошибка сохранения кадра {filename}: {e}")

            QMessageBox.information(self, "Сохранение завершено", f"Сохранено {saved_count} из {len(self.collected_auto_stop_frames_data)} кадров в папку {base_folder_path}.")
            self.collected_auto_stop_frames_data = []
            self.clear_thumbnails()
            self.save_all_frames_button.setEnabled(False)

    def show_error_message(self, message):
        QMessageBox.critical(self, "Ошибка в потоке обработки", message)
        self.live_video_label.setText("Ошибка. См. сообщение.")
        if self.video_thread and not self.video_thread.isRunning():
            log = self.video_thread.emotion_log if hasattr(self.video_thread, 'emotion_log') else []
            self.processing_finished(log)

    def save_report_xlsx(self):
        if not openpyxl:
            QMessageBox.warning(self, "Ошибка", "Библиотека openpyxl не установлена. Сохранение в XLSX невозможно.")
            return
        if not self.collected_emotion_data_for_report:
            QMessageBox.information(self, "Нет данных", "Нет данных для формирования отчета.")
            return

        default_filename = f"emotion_report_{QDateTime.currentDateTime().toString('yyyyMMdd_HHmm')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчет XLSX", default_filename, "XLSX Files (*.xlsx)")
        if path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Лог Эмоций"
                headers = ['Время (с)', 'Кадр', 'ID лица на кадре', 'ID эмоции',
                           'Эмоция (ENG)', 'Эмоция (RU)', 'Уверенность (%)',
                           'BBox X', 'BBox Y', 'BBox Ширина', 'BBox Высота']
                ws.append(headers)
                for entry in self.collected_emotion_data_for_report:
                    bbox = entry.get('bbox', (None, None, None, None))
                    row = [
                        entry.get('time_s', ''), entry.get('frame', ''), entry.get('face_idx_on_frame', ''),
                        entry.get('emotion_idx', ''), entry.get('emotion_name_eng', ''),
                        entry.get('emotion_name_ru', ''), entry.get('confidence', ''),
                        bbox[0], bbox[1], bbox[2], bbox[3]
                    ]
                    ws.append(row)
                for col_idx, column_cells in enumerate(ws.columns):
                    max_length = 0
                    column_letter = get_column_letter(col_idx + 1)
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                wb.save(path)
                QMessageBox.information(self, "Успех", f"Отчет успешно сохранен в {path}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка сохранения отчета", f"Не удалось сохранить отчет: {e}\n{traceback.format_exc()}")

    def create_emotion_pie_chart(self):
        if not self.collected_emotion_data_for_report:
            QMessageBox.information(self, "Нет данных", "Нет данных для создания диаграммы.")
            return

        emotion_counts = Counter()
        temp_frame_emotion_map = {}
        for entry in self.collected_emotion_data_for_report:
            frame_id = entry['frame']
            if frame_id not in temp_frame_emotion_map:
                 temp_frame_emotion_map[frame_id] = entry['emotion_idx']
        for emotion_idx in temp_frame_emotion_map.values():
            emotion_counts[emotion_idx] +=1
        if not emotion_counts:
            QMessageBox.information(self, "Нет данных", "Не удалось извлечь данные об эмоциях для диаграммы.")
            return

        labels_ru = [EMOTION_DICT_RU.get(idx, "Неизвестно") for idx in emotion_counts.keys()]
        sizes = list(emotion_counts.values())
        colors_for_pie = [EMOTION_COLORS[idx % len(EMOTION_COLORS)] for idx in emotion_counts.keys()]
        fig_pie, ax_pie = plt.subplots(figsize=(10, 6), dpi=100)
        wedges, texts, autotexts = ax_pie.pie(sizes, colors=colors_for_pie, autopct='%1.1f%%',
                                              startangle=90, pctdistance=0.85)
        for autotext in autotexts:
            autotext.set_color('white'); autotext.set_fontsize(8); autotext.set_fontweight('bold')
        ax_pie.axis('equal')
        plt.subplots_adjust(left=0.05, right=0.65, top=0.9, bottom=0.1)
        text_y_start = 0.9
        for i, (idx, count) in enumerate(emotion_counts.items()):
            emotion_name = EMOTION_DICT_RU.get(idx, "Неизвестно")
            percentage = count / sum(sizes) * 100
            color_hex = EMOTION_COLORS[idx % len(EMOTION_COLORS)]
            fig_pie.patches.extend([plt.Rectangle((0.7, text_y_start - i*0.05 - 0.01), 0.02, 0.02,
                                                  facecolor=color_hex, edgecolor='gray',
                                                  transform=fig_pie.transFigure, figure=fig_pie)])
            fig_pie.text(0.73, text_y_start - i*0.05, f'{emotion_name}: {percentage:.1f}% ({count} кадров)',
                         transform=fig_pie.transFigure, fontsize=9)
        if emotion_counts:
            dominant_emotion_idx = emotion_counts.most_common(1)[0][0]
            dominant_emotion_name_ru = EMOTION_DICT_RU.get(dominant_emotion_idx, "Неизвестно")
            title_text = f'Итоговая диаграмма эмоций\nПреобладающая эмоция: {dominant_emotion_name_ru}'
        else: title_text = 'Итоговая диаграмма эмоций'
        fig_pie.suptitle(title_text, fontsize=12, y=0.98)
        default_filename = f"emotion_pie_chart_{QDateTime.currentDateTime().toString('yyyyMMdd_HHmm')}.png"
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить диаграмму", default_filename, "PNG Files (*.png)")
        if path:
            try:
                fig_pie.savefig(path, bbox_inches='tight')
                QMessageBox.information(self, "Успех", f"Диаграмма сохранена в {path}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка сохранения диаграммы", f"Не удалось сохранить: {e}\n{traceback.format_exc()}")
        plt.close(fig_pie)

    @pyqtSlot(str)
    def show_enlarged_thumbnail(self, image_path):
        viewer = EnlargedThumbnailViewer(image_path, self)
        viewer.exec_()

    @pyqtSlot(str, bool)
    def add_thumbnail(self, image_path, is_manual):
        if is_manual and image_path in self.manual_saved_thumb_paths:
            return
        if is_manual:
            self.manual_saved_thumb_paths.append(image_path)

        pixmap = QPixmap(image_path)
        if pixmap.isNull():
            print(f"Не удалось загрузить миниатюру: {image_path}")
            return

        thumb_label = ClickableLabel(image_path)
        thumb_label.setPixmap(pixmap.scaled(180, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        thumb_label.setToolTip(os.path.basename(image_path))
        thumb_label.setFixedSize(180,135)
        thumb_label.setAlignment(Qt.AlignCenter)
        thumb_label.setStyleSheet("border: 1px solid lightgray; margin: 2px;")
        thumb_label.leftClicked.connect(self.show_enlarged_thumbnail)

        self.thumbnails_layout.addWidget(thumb_label)
        self.thumbnails_scroll_area.verticalScrollBar().setValue(
            self.thumbnails_scroll_area.verticalScrollBar().maximum())

    def clear_thumbnails(self):
        while self.thumbnails_layout.count():
            child = self.thumbnails_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.manual_saved_thumb_paths = []

        if os.path.exists(self.temp_auto_thumbs_dir):
            try:
                for item in os.listdir(self.temp_auto_thumbs_dir):
                    item_path = os.path.join(self.temp_auto_thumbs_dir, item)
                    try:
                        if os.path.isfile(item_path) or os.path.islink(item_path):
                            os.unlink(item_path)
                        elif os.path.isdir(item_path):
                             shutil.rmtree(item_path)
                    except Exception as e:
                        print(f'Failed to delete {item_path}. Reason: {e}')
            except Exception as e:
                print(f"Error clearing temp thumbnails directory {self.temp_auto_thumbs_dir}: {e}")


    def convert_cv_qt(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        qt_image = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        return QPixmap.fromImage(qt_image)

    @pyqtSlot(int)
    def set_video_duration(self, duration_ms):
        self.current_video_duration_ms = duration_ms
        if duration_ms > 0:
            self.video_slider.setRange(0, duration_ms)
            self.video_slider.setEnabled(True)
            self.total_duration_str = self.format_ms_to_time(duration_ms)
            current_time_str = self.format_ms_to_time(self.video_slider.value())
            self.time_display_label.setText(f"{current_time_str} / {self.total_duration_str}")
        else:
            self.video_slider.setRange(0, 0)
            self.video_slider.setEnabled(False)
            self.total_duration_str = "00:00"
            self.time_display_label.setText(f"00:00 / {self.total_duration_str}")

    @pyqtSlot(int)
    def update_slider_position(self, current_msec_from_video):
        if not self.user_is_scrubbing and self.video_slider.isEnabled():
            self.video_slider.setValue(current_msec_from_video)
        current_time_str = self.format_ms_to_time(current_msec_from_video)
        self.time_display_label.setText(f"{current_time_str} / {self.total_duration_str}")

    @pyqtSlot()
    def handle_clear_plot_request(self):
        self.init_plot()

    def slider_pressed(self):
        if self.video_file_radio.isChecked() and self.video_thread and self.video_thread.isRunning():
            self.user_is_scrubbing = True

    def slider_released(self):
        if self.user_is_scrubbing:
            self.user_is_scrubbing = False
            if self.video_thread and self.video_thread.isRunning() and self.video_file_radio.isChecked():
                new_position_msec = self.video_slider.value()
                self.video_thread.seek_video(new_position_msec)

    def update_time_label_from_slider(self, value_ms):
        current_time_str = self.format_ms_to_time(value_ms)
        self.time_display_label.setText(f"{current_time_str} / {self.total_duration_str}")

    def closeEvent(self, event):
        if self.video_thread and self.video_thread.isRunning():
            self.video_thread.stop_stream()
        if os.path.exists(self.temp_auto_thumbs_dir):
            try:
                shutil.rmtree(self.temp_auto_thumbs_dir)
            except Exception as e:
                print(f"Error removing temp directory on close: {e}")
        event.accept()

if __name__ == "__main__":
    os.makedirs("results", exist_ok=True)
    haar_file = 'haarcascade_frontalface_default.xml'
    haar_path_cv = cv2.data.haarcascades + haar_file
    base_path_exe = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    haar_path_local = os.path.join(base_path_exe, haar_file)
    if not os.path.exists(haar_path_cv) and not os.path.exists(haar_path_local):
        pass
    app = QApplication(sys.argv)
    main_win = EmotionDetectorApp()
    main_win.show()
    sys.exit(app.exec_())